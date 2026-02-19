#!/usr/bin/env python3
import json
import sys
from pathlib import Path
from datetime import datetime, timedelta

import openpyxl


def main():
    payload = json.loads(sys.argv[1])
    xlsx_path = Path(sys.argv[2])
    missing = (payload.get("requirements_extract", {}).get("missing_information", []) or [])[:12]

    dedup = []
    seen = set()
    for item in missing:
      q = (item.get("question") or "").strip()
      if not q or q in seen:
        continue
      seen.add(q)
      dedup.append(item)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    due = (datetime.now() + timedelta(days=2)).date().isoformat()

    start = 2
    for i, m in enumerate(dedup, start=start):
      ev = m.get("evidence", {})
      ws.cell(row=i, column=1).value = m.get("question", "")
      ws.cell(row=i, column=2).value = m.get("why", "")
      ws.cell(row=i, column=3).value = f"{ev.get('file','UNKNOWN')} p{ev.get('page','?')}"
      ws.cell(row=i, column=4).value = due
      ws.cell(row=i, column=5).value = "Draft"

    wb.save(xlsx_path)
    print(json.dumps({"rfi_count": len(dedup)}))


if __name__ == "__main__":
    main()
